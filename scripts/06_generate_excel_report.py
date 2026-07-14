import os
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR   = os.path.join(BASE_DIR, "data")
OUTPUT_DIR = os.path.join(BASE_DIR, "notebooks")

# ── load data ────────────────────────────────────────────────

for fname in [
    "creators.csv", "campaigns.csv", "campaign_creators.csv",
    "posts.csv", "conversions.csv", "audience_demographics.csv",
]:
    path = os.path.join(DATA_DIR, fname)
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"{fname} not found in data/. Run generator scripts first."
        )

creators          = pd.read_csv(os.path.join(DATA_DIR, "creators.csv"))
campaigns         = pd.read_csv(os.path.join(DATA_DIR, "campaigns.csv"))
campaign_creators = pd.read_csv(os.path.join(DATA_DIR, "campaign_creators.csv"))
posts             = pd.read_csv(os.path.join(DATA_DIR, "posts.csv"))
conversions       = pd.read_csv(os.path.join(DATA_DIR, "conversions.csv"))
demographics      = pd.read_csv(os.path.join(DATA_DIR, "audience_demographics.csv"))

print("Data loaded successfully.")

# ── hex colours ──────────────────────────────────────────────
# note: openpyxl does not use # before hex codes

GREEN_DARK  = "0F6E56"
GREEN_LIGHT = "E1F5EE"
BLUE_DARK   = "185FA5"
BLUE_LIGHT  = "E6F1FB"
RED_DARK    = "993C1D"
RED_LIGHT   = "FAECE7"
AMBER_LIGHT = "FAEEDA"
AMBER_DARK  = "854F0B"
GREY_HEADER = "2C2C2C"
WHITE       = "FFFFFF"
LIGHT_GREY  = "F5F5F5"

# ── style helpers ────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=WHITE, size=11):
    return Font(bold=bold, color=color, size=size)

def border():
    thin = Side(style="thin", color="DDDDDD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def set_col_widths(ws, widths):
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

def style_cell(cell, bg=WHITE, fg="2C2C2C", bold=False,
               size=10, align="left"):
    cell.fill      = fill(bg)
    cell.font      = Font(bold=bold, color=fg, size=size)
    cell.alignment = center() if align == "center" else left()
    cell.border    = border()

# ════════════════════════════════════════════════════════════
# COMPUTE METRICS
# ════════════════════════════════════════════════════════════

# ROI per partnership using INR columns
cc = campaign_creators.merge(
    creators[["creator_id", "name", "tier", "platform", "niche",
               "authenticity_risk_level"]],
    on="creator_id", how="left",
)
cc["roi_pct"] = (
    (cc["revenue_attributed_inr"] - cc["creator_fee_inr"])
    / cc["creator_fee_inr"].replace(0, np.nan)
) * 100

# creator-level summary
creator_summary = (
    cc.groupby(["creator_id", "name", "tier",
                "platform", "niche", "authenticity_risk_level"])
    .agg(
        total_fee_inr     = ("creator_fee_inr",        "sum"),
        total_revenue_inr = ("revenue_attributed_inr", "sum"),
        num_campaigns     = ("campaign_id",             "nunique"),
    )
    .reset_index()
)
creator_summary["roi_pct"] = (
    (creator_summary["total_revenue_inr"] - creator_summary["total_fee_inr"])
    / creator_summary["total_fee_inr"].replace(0, np.nan)
) * 100
creator_summary = creator_summary.sort_values(
    "roi_pct", ascending=False
).reset_index(drop=True)

# campaign-level summary
camp = campaigns.merge(
    cc.groupby("campaign_id").agg(
        total_fee_inr     = ("creator_fee_inr",        "sum"),
        total_revenue_inr = ("revenue_attributed_inr", "sum"),
        num_creators      = ("creator_id",              "nunique"),
    ).reset_index(),
    on="campaign_id", how="left",
)
camp["roi_pct"] = (
    (camp["total_revenue_inr"] - camp["total_fee_inr"])
    / camp["total_fee_inr"].replace(0, np.nan)
) * 100
camp = camp.sort_values("roi_pct", ascending=False).reset_index(drop=True)

# executive KPIs
total_revenue   = conversions["order_value_inr"].sum()
total_orders    = len(conversions)
avg_order       = conversions["order_value_inr"].mean()

posts["engagement_rate"] = (
    (posts["likes"] + posts["comments"] + posts["shares"])
    / posts["reach"].replace(0, np.nan)
) * 100

top_platform = posts.groupby("platform")["engagement_rate"].mean().idxmax()
top_format   = posts.groupby("content_format")["engagement_rate"].mean().idxmax()
top_city     = conversions.groupby("location")["order_value_inr"].sum().idxmax()
top_creator  = creator_summary.iloc[0]

# tier summary
tier_stats = (
    cc.groupby("tier")
    .agg(
        creators  = ("creator_id",              "nunique"),
        avg_fee   = ("creator_fee_inr",         "mean"),
        avg_rev   = ("revenue_attributed_inr",  "mean"),
        avg_roi   = ("roi_pct",                 "mean"),
        campaigns = ("campaign_id",              "nunique"),
    )
    .reindex(["nano", "micro", "macro", "mega"])
    .fillna(0)
    .reset_index()
)

# demographics — long format, pull age group averages
age_demo = demographics[demographics["demographic_type"] == "age_group"]
avg_age  = (
    age_demo.groupby("demographic_value")["percentage_share"]
    .mean()
    .reindex(["13-17", "18-24", "25-34", "35-44", "45-54", "55+"])
    .round(1)
)

print("Metrics computed.")

# ════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ════════════════════════════════════════════════════════════

wb = Workbook()

# ── Sheet 1: Executive Summary ───────────────────────────────

ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_view.showGridLines = False

# title
ws1.merge_cells("A1:F1")
ws1["A1"]           = "VibeSignal AI — Executive Summary Report"
ws1["A1"].fill      = fill(BLUE_DARK)
ws1["A1"].font      = Font(bold=True, color=WHITE, size=16)
ws1["A1"].alignment = center()
ws1.row_dimensions[1].height = 45

# subtitle
ws1.merge_cells("A2:F2")
ws1["A2"] = (
    f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')} "
    f"| Synthetic data — portfolio/demo only"
)
ws1["A2"].font      = Font(color="888888", size=10, italic=True)
ws1["A2"].alignment = center()
ws1.row_dimensions[2].height = 18

ws1.append([])

# KPI section header
ws1.merge_cells("A4:F4")
ws1["A4"]           = "KEY PERFORMANCE INDICATORS"
ws1["A4"].fill      = fill(GREY_HEADER)
ws1["A4"].font      = Font(bold=True, color=WHITE, size=11)
ws1["A4"].alignment = center()
ws1.row_dimensions[4].height = 28

# KPI column headers
kpi_hdrs = ["Metric", "Value", "", "Metric", "Value", ""]
for col, val in enumerate(kpi_hdrs, start=1):
    cell = ws1.cell(row=5, column=col, value=val)
    style_cell(cell, bg=BLUE_LIGHT, fg=BLUE_DARK,
               bold=True, size=10, align="center")
ws1.row_dimensions[5].height = 22

# KPI rows — all currency in INR
kpi_data = [
    ("Total Synthetic Revenue",  f"₹{total_revenue:,.2f}",
     "", "Total Conversions",    f"{total_orders:,}",          ""),
    ("Avg Order Value (INR)",    f"₹{avg_order:,.2f}",
     "", "Best Platform",        top_platform.title(),          ""),
    ("Best Content Format",      top_format.title(),
     "", "Top City by Revenue",  top_city,                      ""),
    ("Top Creator",              top_creator["name"],
     "", "Top Creator ROI",      f"{top_creator['roi_pct']:.1f}%", ""),
    ("Total Creators",           "200",
     "", "Total Campaigns",      "70",                          ""),
]

for r, row_data in enumerate(kpi_data, start=6):
    ws1.row_dimensions[r].height = 24
    for c, val in enumerate(row_data, start=1):
        cell = ws1.cell(row=r, column=c, value=val)
        cell.border    = border()
        cell.alignment = left()
        if c in (1, 4):
            cell.fill = fill(LIGHT_GREY)
            cell.font = Font(bold=True, color=GREY_HEADER, size=10)
        elif c in (2, 5):
            cell.fill = fill(GREEN_LIGHT)
            cell.font = Font(bold=True, color=GREEN_DARK, size=11)
        else:
            cell.fill = fill(WHITE)

ws1.append([])

# tier breakdown
ws1.merge_cells("A12:F12")
ws1["A12"]           = "PERFORMANCE BY CREATOR TIER (INR)"
ws1["A12"].fill      = fill(GREY_HEADER)
ws1["A12"].font      = Font(bold=True, color=WHITE, size=11)
ws1["A12"].alignment = center()
ws1.row_dimensions[12].height = 28

tier_hdrs = ["Tier", "Creators", "Avg Fee (₹)",
             "Avg Revenue (₹)", "Avg ROI (%)", "Campaigns"]
for col, val in enumerate(tier_hdrs, start=1):
    cell = ws1.cell(row=13, column=col, value=val)
    style_cell(cell, bg=BLUE_LIGHT, fg=BLUE_DARK,
               bold=True, size=10, align="center")
ws1.row_dimensions[13].height = 22

tier_colors = {
    "nano":  GREEN_LIGHT,
    "micro": BLUE_LIGHT,
    "macro": AMBER_LIGHT,
    "mega":  RED_LIGHT,
}

for r, (_, row) in enumerate(tier_stats.iterrows(), start=14):
    ws1.row_dimensions[r].height = 22
    row_vals = [
        str(row["tier"]).title() if pd.notna(row["tier"]) else "N/A",
        int(row["creators"]) if pd.notna(row["creators"]) else 0,
        f"₹{row['avg_fee']:,.0f}" if pd.notna(row["avg_fee"]) else "₹0",
        f"₹{row['avg_rev']:,.0f}" if pd.notna(row["avg_rev"]) else "₹0",
        round(float(row["avg_roi"]), 1) if pd.notna(row["avg_roi"]) else 0,
        int(row["campaigns"]) if pd.notna(row["campaigns"]) else 0,
    ]
    bg = tier_colors.get(row["tier"], WHITE)
    for c, val in enumerate(row_vals, start=1):
        cell = ws1.cell(row=r, column=c, value=val)
        cell.fill      = fill(bg)
        cell.font      = Font(color=GREY_HEADER, size=10)
        cell.alignment = center()
        cell.border    = border()

ws1.append([])

# audience demographics section — from long-format table
ws1.merge_cells("A19:F19")
ws1["A19"]           = "AVERAGE AUDIENCE AGE DISTRIBUTION (%)"
ws1["A19"].fill      = fill(GREY_HEADER)
ws1["A19"].font      = Font(bold=True, color=WHITE, size=11)
ws1["A19"].alignment = center()
ws1.row_dimensions[19].height = 28

age_hdrs = list(avg_age.index)
for col, val in enumerate(age_hdrs, start=1):
    cell = ws1.cell(row=20, column=col, value=val)
    style_cell(cell, bg=BLUE_LIGHT, fg=BLUE_DARK,
               bold=True, size=10, align="center")

age_vals = list(avg_age.values)
for col, val in enumerate(age_vals, start=1):
    cell = ws1.cell(row=21, column=col, value=round(float(val), 1))
    style_cell(cell, bg=GREEN_LIGHT, fg=GREEN_DARK,
               bold=True, size=11, align="center")

set_col_widths(ws1, [22, 14, 14, 16, 14, 12])

# ── Sheet 2: Creator Performance ─────────────────────────────

ws2 = wb.create_sheet("Creator Performance")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:I1")
ws2["A1"]           = "Creator Performance — Ranked by ROI (INR)"
ws2["A1"].fill      = fill(GREEN_DARK)
ws2["A1"].font      = Font(bold=True, color=WHITE, size=14)
ws2["A1"].alignment = center()
ws2.row_dimensions[1].height = 40

hdrs2 = [
    "Rank", "Creator Name", "Tier", "Platform", "Niche",
    "Risk Level", "Total Fee (₹)", "Total Revenue (₹)", "ROI (%)",
]
for col, val in enumerate(hdrs2, start=1):
    cell = ws2.cell(row=2, column=col, value=val)
    style_cell(cell, bg=GREY_HEADER, fg=WHITE,
               bold=True, size=10, align="center")
ws2.row_dimensions[2].height = 25

for r, (_, row) in enumerate(creator_summary.iterrows(), start=3):
    roi = row["roi_pct"]

    if roi >= 150:
        bg, fg = GREEN_LIGHT, GREEN_DARK
    elif roi >= 50:
        bg, fg = BLUE_LIGHT, BLUE_DARK
    elif roi >= 0:
        bg, fg = AMBER_LIGHT, AMBER_DARK
    else:
        bg, fg = RED_LIGHT, RED_DARK

    row_vals = [
        r - 2,
        row["name"],
        row["tier"].title(),
        row["platform"].title(),
        row["niche"].title(),
        row["authenticity_risk_level"].title(),
        f"₹{row['total_fee_inr']:,.0f}",
        f"₹{row['total_revenue_inr']:,.0f}",
        round(roi, 1),
    ]
    ws2.row_dimensions[r].height = 20
    for c, val in enumerate(row_vals, start=1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.fill      = fill(bg)
        cell.font      = Font(color=fg, size=10)
        cell.alignment = center()
        cell.border    = border()

set_col_widths(ws2, [7, 24, 10, 12, 12, 12, 16, 18, 10])

# ── Sheet 3: Campaign Summary ─────────────────────────────────

ws3 = wb.create_sheet("Campaign Summary")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:H1")
ws3["A1"]           = "Campaign Summary — Ranked by ROI (INR)"
ws3["A1"].fill      = fill(AMBER_DARK)
ws3["A1"].font      = Font(bold=True, color=WHITE, size=14)
ws3["A1"].alignment = center()
ws3.row_dimensions[1].height = 40

hdrs3 = [
    "Rank", "Campaign Name", "Objective", "Status",
    "Budget (₹)", "Revenue (₹)", "Creators", "ROI (%)",
]
for col, val in enumerate(hdrs3, start=1):
    cell = ws3.cell(row=2, column=col, value=val)
    style_cell(cell, bg=GREY_HEADER, fg=WHITE,
               bold=True, size=10, align="center")
ws3.row_dimensions[2].height = 25

for r, (_, row) in enumerate(camp.iterrows(), start=3):
    roi = row["roi_pct"] if not pd.isna(row["roi_pct"]) else 0

    if roi >= 150:
        bg = GREEN_LIGHT
    elif roi >= 50:
        bg = BLUE_LIGHT
    elif roi >= 0:
        bg = AMBER_LIGHT
    else:
        bg = RED_LIGHT

    row_vals = [
        r - 2,
        row["campaign_name"],
        row["objective"].title() if not pd.isna(row["objective"]) else "N/A",
        row["status"].title()    if not pd.isna(row["status"])    else "N/A",
        f"₹{row['total_budget_inr']:,.0f}",
        f"₹{row['total_revenue_inr']:,.0f}" if not pd.isna(row["total_revenue_inr"]) else "₹0",
        int(row["num_creators"])              if not pd.isna(row["num_creators"])     else 0,
        round(roi, 1),
    ]
    ws3.row_dimensions[r].height = 20
    for c, val in enumerate(row_vals, start=1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.fill      = fill(bg)
        cell.font      = Font(color=GREY_HEADER, size=10)
        cell.alignment = center()
        cell.border    = border()

set_col_widths(ws3, [7, 32, 14, 12, 14, 14, 10, 10])

# ── Save ─────────────────────────────────────────────────────

os.makedirs(OUTPUT_DIR, exist_ok=True)

filename = f"VibeSignal_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
out_path = os.path.join(OUTPUT_DIR, filename)
wb.save(out_path)

print()
print("=" * 50)
print("report saved:", filename)
print("location:", out_path)
print()
print("sheets:")
print("  1. Executive Summary   — KPIs, tier breakdown, audience age")
print("  2. Creator Performance — 200 creators ranked by ROI")
print("  3. Campaign Summary    — 70 campaigns ranked by ROI")
print()
print("colour coding:")
print("  green  — ROI above 150%  (scale)")
print("  blue   — ROI 50-150%     (solid)")
print("  amber  — ROI 0-50%       (watch)")
print("  red    — ROI negative    (stop)")
print()